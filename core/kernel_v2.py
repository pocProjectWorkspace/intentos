"""
IntentOS Intent Kernel v2.0.0 — Production Execution Loop.

Fully integrated with Phase 2/3 modules:
  - Security pipeline (input/output scanning, credential leak detection)
  - Inference service (privacy-aware LLM routing, cost tracking)
  - Orchestration (SOP engine, agent scheduler, mode router, message bus)
  - RAG context assembler (file index, task history, experience)

The v1 kernel (kernel.py) is preserved as legacy fallback.
"""

from __future__ import annotations

import importlib
import json
import os
import sys
import time
import uuid
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional

# Ensure project root is on sys.path
if getattr(sys, 'frozen', False):
    _PROJECT_ROOT = sys._MEIPASS
else:
    _PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

# --- Phase 2/3 module imports ---
from core.security.pipeline import SecurityPipeline
from core.security.credential_provider import CredentialProvider, get_api_key
from core.inference.llm import LLMService, LLMResponse
from core.inference.router import PrivacyMode
from core.orchestration.sop import SOPExecutor, Phase, PhaseResult, RecoveryAction, SOPResult
from core.orchestration.scheduler import AgentScheduler, AgentManifest, ExecutionResult
from core.orchestration.mode_router import ModeRouter, ReactMode
from core.orchestration.cost_manager import CostManager
from core.orchestration.message_bus import MessageBus, Message
from core.rag.context import ContextAssembler

# ---------------------------------------------------------------------------
# System prompt (carried forward from v1 kernel)
# ---------------------------------------------------------------------------

SYSTEM_PROMPT = """\
You are the IntentOS Intent Kernel. Your sole job is to receive a natural \
language instruction from the user and return a structured intent object as JSON.

You must respond with ONLY valid JSON — no markdown, no explanation, no wrapping.

The intent object schema:

{
  "raw_input": "<the exact instruction the user typed>",
  "intent": "<category.action in dot notation, e.g. file.rename, image.background_remove, browser.search>",
  "subtasks": [
    {
      "id": "1",
      "agent": "<which agent handles this step, e.g. file_agent, browser_agent, image_agent, media_agent, system_agent, document_agent>",
      "action": "<the specific action the agent should perform>",
      "params": { "<action-specific parameters>" }
    }
  ]
}

Available agents and their actions:

file_agent:
  - list_files: params {path, extension, recursive} — list files in a folder, filter by extension
  - find_files: params {path, pattern, extension, size_gt, size_lt, modified_after} — search files by name/type/size
  - get_metadata: params {path} — file size, dates, type
  - get_disk_usage: params {path} — disk usage summary, largest files
  - rename_file: params {path, new_name}
  - move_file: params {source, destination}
  - copy_file: params {source, destination}
  - create_folder: params {path}
  - delete_file: params {path}
  - read_file: params {path}
  - organize_by_type: params {path} — sort files into subfolders by type
  - bulk_rename: params {path, pattern, replacement}
  - deduplicate: params {path}

system_agent:
  - get_current_date: params {format}

browser_agent:
  - search_web: params {query, max_results}
  - fetch_page: params {url}
  - extract_data: params {url, description, content}

document_agent:
  - create_document: params {filename, content, title, save_path}
  - append_content: params {path, content, heading}
  - read_document: params {path}
  - convert_document: params {path, format}
  - save_document: params {path, destination, filename}

image_agent:
  - remove_background: params {path}
  - resize: params {path, width, height}

media_agent:
  - get_info: params {path} — get media file metadata (duration, codec, bitrate)
  - convert: params {path, format}
  - trim: params {path, start, end}
  - extract_audio: params {path} — extract audio track from video
  - compress: params {path, quality}

Rules:
1. "raw_input" is always the exact text the user typed, unmodified.
2. "intent" is a short dot-notation label: <category>.<action>.
3. "subtasks" is an ordered array with id, agent, action, and params.
4. For multi-step tasks, decompose into minimal ordered subtasks. \
Reference earlier results as "{{1.result}}".
5. If the instruction is ambiguous, pick the most likely interpretation.
6. Always return valid JSON. Never return markdown or explanation text.
7. Use ~ for home directory paths. The accessible folders are: ~/Downloads, \
~/Documents, ~/Desktop. If the user says "my machine" or "my computer", \
search these three folders using one subtask per folder (up to 3 subtasks). \
If the user says "my downloads" or "downloads folder", use ~/Downloads.
8. ONLY use actions listed above. Never invent new actions. \
If the action you want does not exist in the list above, do NOT use it. \
There is no "count" action — file_agent.find_files already returns the count \
of matching files along with the file list. \
To count or find files by type (audio, video, images, etc.), use \
file_agent.find_files with the extension param set to the category name. \
Example: extension "audio" finds all audio files, extension "video" finds \
all video files. You can also use specific extensions like "mp3" or "pdf". \
To find large files, use file_agent.get_disk_usage. \
Keep subtasks minimal — usually 1 subtask is enough. Never use 2 subtasks \
when 1 subtask can do the job.
"""

# ---------------------------------------------------------------------------
# Agent registry
# ---------------------------------------------------------------------------

AGENT_REGISTRY: Dict[str, str] = {
    "file_agent": "capabilities.file_agent.agent",
    "system_agent": "capabilities.system_agent.agent",
    "browser_agent": "capabilities.browser_agent.agent",
    "document_agent": "capabilities.document_agent.agent",
    "image_agent": "capabilities.image_agent.agent",
    "media_agent": "capabilities.media_agent.agent",
}

# Agent manifests for the scheduler
_AGENT_MANIFESTS: Dict[str, Dict[str, Any]] = {
    "file_agent": {
        "version": "1.0.0",
        "actions": [
            "list_files", "find_files", "rename_file", "move_file",
            "copy_file", "create_folder", "delete_file", "read_file",
            "get_metadata", "get_disk_usage", "organize_by_type",
            "bulk_rename", "cleanup_old_files", "deduplicate",
            "sort_by_date", "flatten_folders", "archive_large_files",
        ],
        "permissions": ["~/Documents", "~/Downloads", "~/Desktop"],
        "sandbox_policy": "WorkspaceWrite",
    },
    "browser_agent": {
        "version": "1.0.0",
        "actions": ["search_web", "fetch_page", "extract_data"],
        "permissions": [],
        "sandbox_policy": "ReadOnly",
    },
    "document_agent": {
        "version": "1.0.0",
        "actions": [
            "create_document", "append_content", "read_document",
            "convert_document", "save_document",
        ],
        "permissions": ["~/Documents", "~/Downloads", "~/Desktop"],
        "sandbox_policy": "WorkspaceWrite",
    },
    "system_agent": {
        "version": "1.0.0",
        "actions": ["get_current_date"],
        "permissions": [],
        "sandbox_policy": "ReadOnly",
    },
    "image_agent": {
        "version": "0.1.0",
        "actions": ["remove_background", "resize"],
        "permissions": ["~/Documents", "~/Downloads", "~/Desktop"],
        "sandbox_policy": "WorkspaceWrite",
    },
    "media_agent": {
        "version": "0.1.0",
        "actions": ["get_info", "convert", "trim", "extract_audio", "compress"],
        "permissions": ["~/Documents", "~/Downloads", "~/Desktop"],
        "sandbox_policy": "WorkspaceWrite",
    },
}


# ---------------------------------------------------------------------------
# TaskResult
# ---------------------------------------------------------------------------

@dataclass
class TaskResult:
    """Final result of a process_task() invocation."""
    task_id: str
    status: str  # "success" | "error" | "blocked"
    intent: Optional[Dict[str, Any]] = None
    execution_results: List[ExecutionResult] = field(default_factory=list)
    sop_result: Optional[SOPResult] = None
    report: str = ""
    cost_usd: float = 0.0
    duration_ms: int = 0
    security_warnings: List[str] = field(default_factory=list)
    error: Optional[str] = None


# ---------------------------------------------------------------------------
# Workspace setup
# ---------------------------------------------------------------------------

def _ensure_workspace() -> bool:
    """Ensure ~/.intentos/ exists with full directory structure."""
    home = os.path.expanduser("~")
    base = os.path.join(home, ".intentos")
    first_run = not os.path.exists(base)
    dirs = [
        os.path.join(base, "logs"),
        os.path.join(base, "workspace", "outputs"),
        os.path.join(base, "workspace", "temp"),
        os.path.join(base, "rag"),
    ]
    for d in dirs:
        os.makedirs(d, exist_ok=True)
    audit_file = os.path.join(base, "logs", "audit.jsonl")
    if not os.path.exists(audit_file):
        open(audit_file, "a").close()
    return first_run


# ---------------------------------------------------------------------------
# LLM client wrapper for agent planners
# ---------------------------------------------------------------------------

class _ContentBlock:
    def __init__(self, text: str):
        self.text = text


class _LLMResponse:
    def __init__(self, text: str):
        self.content = [_ContentBlock(text)]
        self.stop_reason = "end_turn"


class _Messages:
    def __init__(self, llm_service):
        self._llm = llm_service

    def create(self, model=None, max_tokens=1024, system=None, messages=None, **kw):
        user_content = ""
        if messages:
            user_content = messages[-1].get("content", "")

        # For planning tasks, prefer the cloud backend with proper
        # system/user separation (critical for structured JSON output).
        cloud = self._llm._router._cloud_backend
        if cloud is not None:
            try:
                result = cloud.generate(
                    user_content,
                    max_tokens=max_tokens,
                    system=system,
                )
                if result and result.text and not result.error:
                    return _LLMResponse(result.text)
            except Exception:
                pass

        # Fallback: concatenate system + user for local models
        if system:
            prompt = f"{system}\n\nUser input: {user_content}\n\nRespond with JSON only."
        else:
            prompt = user_content
        result = self._llm.generate(prompt)
        return _LLMResponse(result.text if result and result.text else "")


class _LLMClientWrapper:
    """Wraps LLMService to look like the Anthropic client API for agent planners."""
    def __init__(self, llm_service):
        self.messages = _Messages(llm_service)


def _build_llm_client(llm_service) -> Optional[_LLMClientWrapper]:
    try:
        backend = llm_service._router._local_backend or llm_service._router._cloud_backend
        if backend:
            return _LLMClientWrapper(llm_service)
    except Exception:
        pass
    return None


# ---------------------------------------------------------------------------
# IntentKernel
# ---------------------------------------------------------------------------

class IntentKernel:
    """The v2 kernel — fully integrated with all Phase 2/3 modules."""

    VERSION = "2.0.0"

    def __init__(
        self,
        privacy_mode: PrivacyMode = PrivacyMode.SMART_ROUTING,
        budget: Optional[float] = None,
    ) -> None:
        # Workspace
        _ensure_workspace()
        self._workspace = os.path.join(os.path.expanduser("~"), ".intentos", "workspace")

        # Security
        self.security = SecurityPipeline(strict_mode=True)
        self.credentials = CredentialProvider()

        # Inference
        self.llm = LLMService(
            privacy_mode=privacy_mode,
            credential_provider=self.credentials,
            budget=budget,
        )

        # Orchestration
        self.scheduler = AgentScheduler(workspace=self._workspace)
        self.mode_router = ModeRouter()
        self.message_bus = MessageBus()
        self.sop = None  # Created per-task

        # Context
        self.context_assembler = ContextAssembler()

        # Task history (in-memory for the session)
        self._task_history: List[TaskResult] = []

        # Register all agents
        self._register_agents()

    # -- Agent registration -------------------------------------------------

    def _register_agents(self) -> None:
        """Register all capability agents with the scheduler."""
        for agent_name, module_path in AGENT_REGISTRY.items():
            manifest_data = _AGENT_MANIFESTS.get(agent_name, {})
            home = os.path.expanduser("~")
            permissions = [
                p.replace("~", home) for p in manifest_data.get("permissions", [])
            ]
            permissions.append(self._workspace)

            manifest = AgentManifest(
                name=agent_name,
                version=manifest_data.get("version", "0.0.0"),
                actions=manifest_data.get("actions", []),
                permissions=permissions,
                sandbox_policy=manifest_data.get("sandbox_policy", "WorkspaceWrite"),
            )

            # Create a handler that dynamically loads and calls the agent module
            handler = _make_agent_handler(agent_name, module_path)
            self.scheduler.register_agent(agent_name, handler, manifest)

    # -- Main execution loop ------------------------------------------------

    def process_task(self, user_input: str) -> TaskResult:
        """The main execution loop — a single task from input to result."""
        task_id = str(uuid.uuid4())
        task_start = time.monotonic()
        cost_before = self.llm.get_total_spent()

        # 1. Security: scan input
        input_scan = self.security.process_input(user_input)
        if input_scan.blocked:
            return TaskResult(
                task_id=task_id,
                status="blocked",
                error="Input blocked by security pipeline",
                security_warnings=input_scan.warnings,
                duration_ms=int((time.monotonic() - task_start) * 1000),
            )

        sanitized_input = input_scan.sanitized_input

        # 2. Context: assemble relevant context from RAG
        assembled = self.context_assembler.build_context(sanitized_input)

        # 3. Create SOP executor for this task
        sop = SOPExecutor(context={"task_id": task_id, "user_input": sanitized_input})
        self.sop = sop

        # Register phase handlers
        sop.register_handler(Phase.PARSE, lambda inp, ctx: self._phase_parse(inp, ctx))
        sop.register_handler(Phase.PLAN, lambda inp, ctx: self._phase_plan(inp, ctx))
        sop.register_handler(Phase.VALIDATE, lambda inp, ctx: self._phase_validate(inp, ctx))
        sop.register_handler(Phase.PREVIEW, lambda inp, ctx: self._phase_preview(inp, ctx))
        sop.register_handler(Phase.EXECUTE, lambda inp, ctx: self._phase_execute(inp, ctx))
        sop.register_handler(Phase.VERIFY, lambda inp, ctx: self._phase_verify(inp, ctx))
        sop.register_handler(Phase.REPORT, lambda inp, ctx: self._phase_report(inp, ctx))

        # Register error handler
        sop.register_error_handler(self._handle_sop_error)

        # Determine if preview is needed (will be checked inside SOP run)
        # For now, run all phases; PREVIEW handler itself decides whether to
        # do real work or pass through.
        sop_result = sop.run(
            initial_input={"user_input": sanitized_input, "context": assembled},
            skip_preview=False,
        )

        duration_ms = int((time.monotonic() - task_start) * 1000)
        cost_after = self.llm.get_total_spent()
        task_cost = cost_after - cost_before

        # Extract results from SOP
        intent = None
        execution_results = []
        report = ""

        parse_phase = sop_result.get_phase(Phase.PARSE)
        if parse_phase and parse_phase.output:
            intent = parse_phase.output

        execute_phase = sop_result.get_phase(Phase.EXECUTE)
        if execute_phase and execute_phase.output:
            execution_results = execute_phase.output if isinstance(execute_phase.output, list) else []

        report_phase = sop_result.get_phase(Phase.REPORT)
        if report_phase and report_phase.output:
            report = report_phase.output if isinstance(report_phase.output, str) else str(report_phase.output)

        # Build task result
        task_result = TaskResult(
            task_id=task_id,
            status=sop_result.overall_status,
            intent=intent,
            execution_results=execution_results,
            sop_result=sop_result,
            report=report,
            cost_usd=task_cost,
            duration_ms=duration_ms,
            security_warnings=input_scan.warnings if input_scan.has_warnings else [],
        )

        # Record in history
        self._task_history.append(task_result)

        # Record in RAG context assembler
        try:
            agents_used = list({r.agent_name for r in execution_results if hasattr(r, "agent_name")})
            files_accessed = []
            for r in execution_results:
                if hasattr(r, "paths_accessed"):
                    files_accessed.extend(r.paths_accessed)
            params = intent.get("subtasks", [{}])[0].get("params", {}) if intent else {}
            self.context_assembler.record_task(
                raw_input=user_input,
                intent=intent.get("intent", "unknown") if intent else "unknown",
                agents=agents_used,
                files=files_accessed or [""],
                params=params,
                status=sop_result.overall_status,
                duration=duration_ms,
            )
        except Exception:
            pass  # Context recording is best-effort

        # Publish completion message on the bus
        self.message_bus.publish(Message(
            content=f"Task {task_id} completed: {sop_result.overall_status}",
            cause_by="task_complete",
            sent_from="kernel",
            payload={"task_id": task_id, "status": sop_result.overall_status},
        ))

        return task_result

    # -- File-aware processing ----------------------------------------------

    def process_task_with_file(
        self, user_input: str, file_path: str, file_name: str, mime: str,
    ) -> TaskResult:
        """Process a task with an attached file (document or image).

        For images: sends to Gemma 4 multimodal API for visual analysis.
        For documents: extracts text, then asks the LLM to respond.
        """
        task_id = str(uuid.uuid4())
        task_start = time.monotonic()
        cost_before = self.llm.get_total_spent()

        try:
            is_image = mime.startswith("image/")

            if is_image:
                response_text = self._process_image(user_input, file_path, file_name)
            else:
                response_text = self._process_document(user_input, file_path, file_name, mime)

        except Exception as exc:
            return TaskResult(
                task_id=task_id,
                status="error",
                error=str(exc),
                duration_ms=int((time.monotonic() - task_start) * 1000),
            )

        duration_ms = int((time.monotonic() - task_start) * 1000)
        cost_after = self.llm.get_total_spent()

        return TaskResult(
            task_id=task_id,
            status="success",
            report=response_text,
            cost_usd=cost_after - cost_before,
            duration_ms=duration_ms,
        )

    def _process_image(self, user_input: str, file_path: str, file_name: str) -> str:
        """Send an image to the multimodal LLM for analysis."""
        import base64

        with open(file_path, "rb") as f:
            image_b64 = base64.b64encode(f.read()).decode("ascii")

        prompt = (
            f"The user uploaded an image called '{file_name}'. "
            f"Their request: {user_input}\n\n"
            "Analyze the image and respond to their request. "
            "Be concise and helpful. Never use technical jargon."
        )

        # Try local multimodal first (Gemma 4 supports images)
        backend = self.llm._router._local_backend
        if backend is not None and hasattr(backend, "generate"):
            result = backend.generate(prompt, images=[image_b64])
            if result.text and not result.error:
                return result.text.strip()

        # Fallback to cloud (without image — describe what we see)
        result = self.llm.generate(prompt)
        if result.text:
            return result.text.strip()

        return "I received the image but couldn't analyze it. Try again or check your AI backend."

    def _process_document(
        self, user_input: str, file_path: str, file_name: str, mime: str,
    ) -> str:
        """Extract text from a document and send to LLM for analysis."""
        text = self._extract_text(file_path, file_name, mime)

        if not text.strip():
            return f"I couldn't extract any text from '{file_name}'. The file may be empty or in an unsupported format."

        # Truncate to ~8000 chars to fit context window
        max_chars = 8000
        truncated = len(text) > max_chars
        doc_text = text[:max_chars]

        prompt = (
            f"The user uploaded a document called '{file_name}'. "
            f"Their request: {user_input}\n\n"
            f"--- DOCUMENT CONTENT ---\n{doc_text}\n--- END ---"
        )
        if truncated:
            prompt += f"\n(Document truncated — showing first {max_chars} of {len(text)} characters)"

        prompt += (
            "\n\nRespond to the user's request based on the document content. "
            "Be concise and helpful. Never use technical jargon."
        )

        result = self.llm.generate(prompt)
        if result.text:
            return result.text.strip()

        return f"I read '{file_name}' but couldn't generate a response. Try again."

    @staticmethod
    def _extract_text(file_path: str, file_name: str, mime: str) -> str:
        """Extract text content from various document formats."""
        ext = os.path.splitext(file_name)[1].lower()

        # Plain text formats
        if mime.startswith("text/") or ext in (".txt", ".csv", ".md", ".json", ".log"):
            with open(file_path, "r", errors="replace") as f:
                return f.read()

        # PDF
        if ext == ".pdf" or mime == "application/pdf":
            try:
                from pypdf import PdfReader
                reader = PdfReader(file_path)
                pages = []
                for page in reader.pages:
                    text = page.extract_text()
                    if text:
                        pages.append(text)
                return "\n\n".join(pages)
            except ImportError:
                return "[PDF reading requires pypdf — already installed]"
            except Exception as e:
                return f"[Could not read PDF: {e}]"

        # DOCX
        if ext == ".docx" or "wordprocessingml" in mime:
            try:
                from docx import Document
                doc = Document(file_path)
                return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
            except ImportError:
                return "[DOCX reading requires python-docx — already installed]"
            except Exception as e:
                return f"[Could not read DOCX: {e}]"

        # XLSX (read as CSV-like text)
        if ext == ".xlsx" or "spreadsheetml" in mime:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(file_path, read_only=True, data_only=True)
                lines = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    lines.append(f"--- Sheet: {sheet} ---")
                    for row in ws.iter_rows(values_only=True):
                        cells = [str(c) if c is not None else "" for c in row]
                        lines.append("\t".join(cells))
                return "\n".join(lines)
            except ImportError:
                return "[XLSX reading requires openpyxl]"
            except Exception as e:
                return f"[Could not read XLSX: {e}]"

        return f"[Unsupported document format: {ext}]"

    # -- Conversational chat (no agent routing) -----------------------------

    def process_chat(self, user_input: str, history: list) -> TaskResult:
        """Process a follow-up message using conversation history + semantic memory.

        This bypasses the agent routing pipeline and sends the full
        conversation context to the LLM for a direct response. Also queries
        the RAG system for relevant past conversations and files.
        """
        task_id = str(uuid.uuid4())
        task_start = time.monotonic()
        cost_before = self.llm.get_total_spent()

        try:
            # Build conversation context from history
            conv_lines = []
            for msg in history[-20:]:  # last 20 messages for context
                role = msg.get("role", "user")
                content = msg.get("content", "")
                if role == "user":
                    conv_lines.append(f"User: {content}")
                else:
                    conv_lines.append(f"Assistant: {content[:800]}")

            context_str = "\n\n".join(conv_lines)

            # Query semantic memory for relevant context
            memory_context = ""
            try:
                assembled = self.context_assembler.build_context(user_input)
                if assembled and assembled.context_text:
                    memory_context = assembled.context_text
            except Exception:
                pass

            memory_block = ""
            if memory_context:
                memory_block = (
                    f"\n\nRelevant memory from past sessions:\n{memory_context}\n"
                )

            prompt = (
                "You are IntentOS, a helpful AI assistant. Continue this conversation "
                "naturally. Respond directly to the user's latest message.\n\n"
                "Rules:\n"
                "- Respond with the content directly in your message\n"
                "- Do NOT say you will create or save a file — put the content right here\n"
                "- Use markdown formatting for structure (headings, lists, bold)\n"
                "- Be concise and helpful\n"
                "- If asked to draft/write/create something, write it inline\n"
                "- If the user asks about past work, files, or conversations, use the "
                "memory context below to answer accurately\n"
                "- If you find matching files or past conversations in memory, mention "
                "the specific file names and paths\n\n"
                f"Conversation so far:\n{context_str}\n"
                f"{memory_block}\n"
                f"User: {user_input}\n\n"
                "Assistant:"
            )

            result = self.llm.generate(prompt)
            response_text = result.text.strip() if result and result.text else "I couldn't generate a response."

        except Exception as exc:
            return TaskResult(
                task_id=task_id,
                status="error",
                error=str(exc),
                duration_ms=int((time.monotonic() - task_start) * 1000),
            )

        duration_ms = int((time.monotonic() - task_start) * 1000)
        cost_after = self.llm.get_total_spent()

        return TaskResult(
            task_id=task_id,
            status="success",
            report=response_text,
            cost_usd=cost_after - cost_before,
            duration_ms=duration_ms,
        )

    # -- SOP Phase Handlers -------------------------------------------------

    def _phase_parse(self, initial_input: Any, context: Dict) -> Dict:
        """PARSE: Use LLM to parse intent from user input + assembled context."""
        user_input = initial_input["user_input"]
        assembled = initial_input.get("context")

        # Build prompt with context
        context_text = assembled.context_text if assembled else ""
        if context_text:
            prompt = f"{context_text}\n\nUser instruction: {user_input}"
        else:
            prompt = user_input

        result = self.llm.parse_intent(prompt, SYSTEM_PROMPT)
        if result is None:
            has_local = self.llm._router._local_backend is not None
            has_cloud = self.llm._router._cloud_backend is not None
            config = self.llm.get_config()
            model = config.get("local_model", "unknown")

            if not has_local and not has_cloud:
                raise ValueError(
                    "No AI backend is available. Install Ollama "
                    f"(https://ollama.com/download) and run: ollama pull {model} "
                    "— or set ANTHROPIC_API_KEY in your .env file."
                )
            raise ValueError(
                f"Could not understand that instruction — the local model ({model}) "
                f"did not return a usable response. Try rephrasing, or check that "
                f"Ollama is running and the model is pulled (ollama pull {model})"
            )

        # Ensure raw_input is preserved
        if "raw_input" not in result:
            result["raw_input"] = user_input

        return result

    def _phase_plan(self, intent: Dict, context: Dict) -> Dict:
        """PLAN: Decompose intent into subtasks, select execution mode."""
        subtasks = intent.get("subtasks", [])
        if not subtasks:
            raise ValueError("No subtasks generated — try rephrasing the instruction")

        mode = self.mode_router.select_mode(intent.get("raw_input", ""), subtasks)
        return {"intent": intent, "subtasks": subtasks, "mode": mode}

    def _phase_validate(self, plan: Dict, context: Dict) -> Dict:
        """VALIDATE: Check all agents exist, permissions valid."""
        for subtask in plan["subtasks"]:
            agent = subtask.get("agent")
            if not self.scheduler.is_registered(agent):
                raise ValueError(f"Agent '{agent}' is not available")
        return plan

    def _phase_preview(self, plan: Dict, context: Dict) -> Dict:
        """PREVIEW: Dry-run for destructive/batch operations."""
        subtasks = plan["subtasks"]
        needs_preview = SOPExecutor.needs_preview(subtasks)

        if not needs_preview:
            return {"plan": plan, "preview": None, "skipped": True}

        dry_context = {**self._build_context(), "dry_run": True}
        results = self.scheduler.execute_sequential(subtasks, dry_context)
        return {"plan": plan, "preview": results, "skipped": False}

    def _phase_execute(self, preview_data: Dict, context: Dict) -> List[ExecutionResult]:
        """EXECUTE: Run subtasks through scheduler with security pipeline."""
        plan = preview_data.get("plan", preview_data)
        real_context = self._build_context()

        mode = plan.get("mode", ReactMode.BY_ORDER)
        subtasks = plan["subtasks"]

        if mode == ReactMode.BY_ORDER:
            results = self.scheduler.execute_sequential(subtasks, real_context)
        elif mode == ReactMode.PLAN_AND_ACT:
            results = self.scheduler.execute_sequential(subtasks, real_context)
        else:  # REACT — sequential for now (dynamic selection is Phase 4)
            results = self.scheduler.execute_sequential(subtasks, real_context)

        return results

    def _phase_verify(self, results: List[ExecutionResult], context: Dict) -> List[ExecutionResult]:
        """VERIFY: Check results match intent, scan for leaks."""
        for result in results:
            if result.output:
                scan = self.security.process_output(result.output)
                if scan.had_leaks:
                    result.output = scan.sanitized_output
        return results

    def _phase_report(self, results: List[ExecutionResult], context: Dict) -> str:
        """REPORT: Generate a conversational response from execution results."""
        raw_input = context.get("user_input", "")
        return self._generate_response(raw_input, results)

    # -- Error handling -----------------------------------------------------

    def _handle_sop_error(self, phase_result: PhaseResult) -> RecoveryAction:
        """Decide recovery action on phase failure."""
        phase = phase_result.phase
        error = phase_result.error_message or ""

        # Parse failures can be retried once (LLM might return valid JSON)
        if phase == Phase.PARSE:
            return RecoveryAction.ABORT

        # Validation failure: abort (can't recover from missing agents)
        if phase == Phase.VALIDATE:
            return RecoveryAction.ABORT

        # Execution failure: abort for now (retry logic is Phase 4)
        if phase == Phase.EXECUTE:
            return RecoveryAction.ABORT

        return RecoveryAction.ABORT

    # -- Context builder ----------------------------------------------------

    def _build_context(self) -> Dict:
        """Build the context dict injected into every agent call."""
        home = os.path.expanduser("~")

        # Create an LLM client wrapper that agents can use for planning.
        # The file_agent planner calls client.messages.create(model, system, messages)
        # — we translate that to our backend's generate() method.
        llm_client = _build_llm_client(self.llm)

        return {
            "user": os.getenv("USER", "unknown"),
            "workspace": self._workspace,
            "granted_paths": [
                os.path.join(home, "Documents"),
                os.path.join(home, "Downloads"),
                os.path.join(home, "Desktop"),
                self._workspace,
            ],
            "task_id": str(uuid.uuid4()),
            "dry_run": False,
            "llm_client": llm_client,
        }

    # -- Result formatting --------------------------------------------------

    def _format_results(self, results: List[ExecutionResult]) -> str:
        """Format execution results into a human-readable report."""
        if not results:
            return "No results."

        lines = []
        success_count = 0
        error_count = 0
        skipped_count = 0

        for r in results:
            prefix = ""
            if r.status == "success":
                prefix = "[OK]"
                success_count += 1
            elif r.status == "error":
                prefix = "[ERR]"
                error_count += 1
            elif r.status == "skipped":
                prefix = "[SKIP]"
                skipped_count += 1

            sid = r.subtask_id or "?"
            line = f"  {prefix} [{sid}] {r.agent_name}.{r.action}"
            if r.error:
                line += f" — {r.error}"
            lines.append(line)
            # Show the action_performed summary from successful results
            if r.status == "success" and r.output:
                performed = r.output.get("action_performed")
                if performed:
                    lines.append(f"        → {performed}")

        # Summary
        lines.append("")
        lines.append(f"  Total: {len(results)} | OK: {success_count} | "
                      f"Errors: {error_count} | Skipped: {skipped_count}")

        return "\n".join(lines)

    def _generate_response(self, user_input: str, results: List[ExecutionResult]) -> str:
        """Generate a conversational response from execution results."""
        # Build a compact summary of what happened
        result_summaries = []
        for r in results:
            summary = {"agent": r.agent_name, "action": r.action, "status": r.status}
            if r.output:
                performed = r.output.get("action_performed", "")
                if performed:
                    summary["summary"] = performed
                # Include result data (truncated for large lists)
                result_data = r.output.get("result")
                if isinstance(result_data, list):
                    summary["count"] = len(result_data)
                    summary["items"] = result_data[:10]  # first 10 for context
                    if len(result_data) > 10:
                        summary["truncated"] = True
                elif result_data is not None:
                    summary["result"] = result_data
            if r.error:
                summary["error"] = r.error
            result_summaries.append(summary)

        response_prompt = (
            "You are IntentOS, a friendly AI assistant. The user asked something and "
            "agents executed tasks to fulfill it. Generate a short, natural, conversational "
            "response that directly answers the user's question.\n\n"
            "Rules:\n"
            "- Be concise and friendly — 1-3 sentences max\n"
            "- Answer the question directly (e.g., 'You have 189 PDF files in your Downloads folder.')\n"
            "- If files were found, mention the count and optionally a few notable ones\n"
            "- If something failed, explain in plain language what went wrong\n"
            "- Never use technical jargon (no 'agent', 'subtask', 'execution')\n"
            "- Never use markdown formatting\n"
            "- Do not mention IntentOS or yourself\n\n"
            f"User asked: {user_input}\n\n"
            f"Results: {json.dumps(result_summaries, default=str)}\n\n"
            "Your response:"
        )

        try:
            llm_response = self.llm.generate(response_prompt)
            if llm_response and llm_response.text:
                return llm_response.text.strip()
        except Exception:
            pass

        # Fallback: use action_performed summaries
        fallback_parts = []
        for r in results:
            if r.output and r.output.get("action_performed"):
                fallback_parts.append(r.output["action_performed"])
            elif r.error:
                fallback_parts.append(r.error)
        return "  " + " | ".join(fallback_parts) if fallback_parts else self._format_results(results)

    # -- CLI subcommand handler ---------------------------------------------

    def handle_command(self, command: str) -> str:
        """Handle !-prefixed CLI subcommands. Returns output string."""
        parts = command.strip().split()
        cmd = parts[0].lower() if parts else ""

        if cmd == "status":
            return self._cmd_status()
        elif cmd == "cost":
            return self._cmd_cost()
        elif cmd == "history":
            return self._cmd_history()
        elif cmd == "credentials":
            return self._cmd_credentials()
        elif cmd == "security":
            return self._cmd_security()
        elif cmd == "speak":
            text = " ".join(parts[1:]) if len(parts) > 1 else ""
            return self._cmd_speak(text)
        else:
            return (
                f"Unknown command: {cmd}\n"
                "Available: !status, !cost, !history, !credentials, !security, !speak"
            )

    def _cmd_status(self) -> str:
        """Hardware profile, model, privacy mode, cost summary."""
        config = self.llm.get_config()
        hw = config.get("hardware", {})
        stats = self.llm.get_stats()
        lines = [
            "IntentOS Kernel v2.0.0 — Status",
            "=" * 40,
            f"  Model (local):   {config.get('local_model', 'N/A')}",
            f"  Model (cloud):   {config.get('cloud_model', 'N/A')}",
            f"  Privacy mode:    {config.get('privacy_mode', 'N/A')}",
            f"  Recommended:     {config.get('recommended_model', 'N/A')}",
            "",
            f"  Platform:        {hw.get('platform', 'N/A')}",
            f"  CPU:             {hw.get('cpu_model', 'N/A')} ({hw.get('cpu_cores', '?')} cores)",
            f"  RAM:             {hw.get('ram_gb', '?')} GB",
            f"  GPU:             {hw.get('gpu', {}).get('model', 'None') if hw.get('gpu') else 'None'}",
            "",
            f"  Agents:          {len(self.scheduler.list_agents())} registered",
            f"  Tasks this run:  {len(self._task_history)}",
            f"  Total cost:      ${stats['total_cost_usd']:.4f}",
        ]
        return "\n".join(lines)

    def _cmd_cost(self) -> str:
        """Detailed cost breakdown."""
        report = self.llm.get_cost_report()
        report_dict = report.to_dict()
        lines = [
            "Cost Report",
            "=" * 40,
            f"  Total spent:     ${report_dict['total_spent_usd']:.4f}",
            f"  Input tokens:    {report_dict['total_input_tokens']}",
            f"  Output tokens:   {report_dict['total_output_tokens']}",
            f"  API calls:       {report_dict['call_count']}",
        ]
        budget = self.llm.get_remaining_budget()
        if budget is not None:
            lines.append(f"  Remaining:       ${budget:.4f}")

        if report_dict.get("by_model"):
            lines.append("")
            lines.append("  By model:")
            for model, usage in report_dict["by_model"].items():
                lines.append(
                    f"    {model}: {usage['total_tokens']} tokens, "
                    f"${usage['cost_usd']:.4f}, {usage['call_count']} calls"
                )

        return "\n".join(lines)

    def _cmd_history(self) -> str:
        """Recent task history."""
        if not self._task_history:
            return "No tasks executed this session."

        lines = [
            "Task History",
            "=" * 40,
        ]
        for i, t in enumerate(self._task_history[-10:], 1):
            intent_label = t.intent.get("intent", "?") if t.intent else "?"
            lines.append(
                f"  {i}. [{t.status}] {intent_label} — "
                f"{t.duration_ms}ms, ${t.cost_usd:.4f}"
            )
        return "\n".join(lines)

    def _cmd_credentials(self) -> str:
        """Credential management info."""
        lines = [
            "Credentials",
            "=" * 40,
            f"  Provider:  CredentialProvider",
            f"  Status:    Active",
        ]
        return "\n".join(lines)

    def _cmd_security(self) -> str:
        """Security pipeline stats."""
        stats = self.security.get_stats()
        lines = [
            "Security Pipeline",
            "=" * 40,
            f"  Total scans:       {stats['total_scans']}",
            f"  Inputs blocked:    {stats['inputs_blocked']}",
            f"  Outputs blocked:   {stats['outputs_blocked']}",
            f"  Leaks redacted:    {stats['leaks_redacted']}",
            f"  Policy violations: {stats['policy_violations']}",
        ]
        return "\n".join(lines)

    def _cmd_speak(self, text: str) -> str:
        """Speak text aloud using the TTS engine."""
        if not text:
            return "Usage: !speak <text to speak>"
        try:
            from core.voice.tts import VoiceOutput
            tts = VoiceOutput()
            best = tts.get_best_available()
            result = tts.speak_and_play(text)
            if result:
                return f"  Spoke {result.duration_seconds:.1f}s via {result.provider}"
            return f"  Voice output not available. Best provider: {best.value}"
        except Exception as e:
            return f"  Voice output error: {e}"

    # -- Display helper -----------------------------------------------------

    def display_result(self, result: TaskResult, console=None) -> None:
        """Print a TaskResult to the terminal using rich if available."""
        try:
            from rich.console import Console
            from rich.panel import Panel
            from rich.text import Text
            c = console or Console()
        except ImportError:
            self._display_result_plain(result)
            return

        c.print()

        if result.status == "blocked":
            c.print(Panel(
                result.error or "Blocked by security pipeline",
                title="Blocked",
                border_style="red",
                padding=(0, 2),
            ))
            for w in result.security_warnings:
                c.print(f"  [yellow]Warning:[/yellow] {w}")
            return

        if result.status == "error":
            error_text = ""
            sop = result.sop_result
            if sop:
                for pr in sop.phases:
                    if pr.status == "error":
                        error_text += f"{pr.error_message}\n"
            elif result.error:
                error_text = result.error
            c.print(Panel(
                error_text.strip() or "Something went wrong",
                title="Error",
                border_style="red",
                padding=(0, 2),
            ))
            return

        # Success — conversational response
        report = result.report or "Done."
        c.print(f"  {report}")

        # Compact footer
        secs = result.duration_ms / 1000
        cost = f"${result.cost_usd:.4f}" if result.cost_usd > 0 else "free (local)"
        n_err = sum(1 for r in result.execution_results if r.status == "error")
        footer_parts = [f"{secs:.1f}s", cost]
        if n_err:
            footer_parts.append(f"{n_err} error(s)")
        c.print(f"  [dim]{' · '.join(footer_parts)}[/dim]")
        c.print()

    def _display_result_plain(self, result: TaskResult) -> None:
        """Fallback display without rich."""
        print()
        if result.status == "blocked":
            print(f"  BLOCKED: {result.error or 'security pipeline'}")
            return
        if result.status == "error":
            sop = result.sop_result
            if sop:
                for pr in sop.phases:
                    if pr.status == "error":
                        print(f"  ERROR: {pr.error_message}")
            elif result.error:
                print(f"  ERROR: {result.error}")
            return
        print(f"  {result.report or 'Done.'}")
        secs = result.duration_ms / 1000
        cost = f"${result.cost_usd:.4f}" if result.cost_usd > 0 else "free (local)"
        print(f"  [{secs:.1f}s · {cost}]")
        print()


# ---------------------------------------------------------------------------
# Agent handler factory
# ---------------------------------------------------------------------------

def _make_agent_handler(agent_name: str, module_path: str):
    """Create a handler function that dynamically loads an agent module."""
    def handler(input_dict: Dict, context: Dict) -> Dict:
        try:
            agent_module = importlib.import_module(module_path)
        except ImportError:
            return {
                "status": "error",
                "error": {"code": "AGENT_NOT_AVAILABLE", "message": f"Module '{module_path}' not found"},
                "metadata": {"files_affected": 0},
            }
        agent_input = {
            "action": context.get("action", "unknown"),
            "params": input_dict,
            "context": context,
        }
        try:
            return agent_module.run(agent_input)
        except Exception as exc:
            return {
                "status": "error",
                "error": {"code": "AGENT_CRASH", "message": str(exc)},
                "metadata": {"files_affected": 0},
            }
    return handler


# ---------------------------------------------------------------------------
# Interactive backend setup (when no backend is available)
# ---------------------------------------------------------------------------

def _interactive_setup(kernel: IntentKernel, recommended_model: str) -> IntentKernel:
    """Ask the user how they want to run AI, then set it up for them."""
    import getpass

    print()
    print("  IntentOS needs an AI backend to work. Let's set one up.\n")
    print("    [1] Private  \u2014 Install local AI on your device (free, offline)")
    print("    [2] Cloud    \u2014 Use your API key (Anthropic, OpenAI, etc.)")
    print("    [3] Skip     \u2014 I'll set it up later")
    print()

    choice = input("  Choose [1/2/3] (default: 1): ").strip()

    if choice == "3":
        return kernel

    if choice == "2":
        # Cloud API key setup
        print()
        print("  Which provider?")
        print("    [1] Anthropic (Claude)  \u2014 recommended")
        print("    [2] OpenAI (GPT)")
        print()
        provider_choice = input("  Choose [1/2] (default: 1): ").strip()

        if provider_choice == "2":
            key_name = "OPENAI_API_KEY"
            env_name = "OPENAI_API_KEY"
            print("\n  Get a key at: https://platform.openai.com/api-keys")
        else:
            key_name = "ANTHROPIC_API_KEY"
            env_name = "ANTHROPIC_API_KEY"
            print("\n  Get a key at: https://console.anthropic.com/settings/keys")

        print()
        api_key = getpass.getpass("  Paste your API key: ").strip()

        if api_key:
            # Write to .env
            env_path = os.path.join(_PROJECT_ROOT, ".env")
            with open(env_path, "a") as f:
                f.write(f"\n{env_name}={api_key}\n")
            os.environ[env_name] = api_key
            print("  [ok] Key saved\n")

            # Re-initialize the kernel with the new key
            print("  [..] Connecting...")
            kernel = IntentKernel()
            print("  [ok] Ready\n")
        else:
            print("  No key entered.\n")

        return kernel

    # Default: choice == "1" or empty — Install Ollama locally
    print()
    try:
        from core.inference.ollama_manager import (
            OllamaManager,
            OllamaInstallError,
        )

        manager = OllamaManager()

        # Step 1: Install Ollama if needed
        if not manager.is_installed():
            print("  IntentOS will now install a small local AI engine.")
            print("  This is a one-time setup (~100 MB download).\n")
            proceed = input("  Continue? [Y/n]: ").strip()
            if proceed.lower() == "n":
                return kernel

            print("  [..] Installing local AI engine...")
            try:
                manager.install(silent=True)
                print("  [ok] Installed\n")
            except OllamaInstallError as exc:
                print(f"  [!!] {exc}\n")
                return kernel
        else:
            print("  [ok] Local AI engine found\n")

        # Step 2: Start daemon
        if not manager.is_running():
            print("  [..] Starting local AI engine...")
            if manager.start_daemon():
                print("  [ok] Running\n")
            else:
                print("  [!!] Could not start. Try restarting your computer.\n")
                return kernel

        # Step 3: Pull models
        print(f"  [..] Downloading AI model ({recommended_model})...")
        print("       This takes a few minutes on first run.\n")

        def _show_progress(progress):
            if progress.status == "complete":
                print(f"  [ok] {progress.model}")
            elif progress.total_gb > 0:
                bar_w = 20
                filled = int(bar_w * progress.percent / 100)
                bar = "\u2588" * filled + "\u2591" * (bar_w - filled)
                print(f"\r  [{bar}] {progress.percent:5.1f}%", end="", flush=True)
            else:
                print(f"\r  {progress.status}...", end="", flush=True)

        result = manager.setup_for_local(recommended_model, _show_progress)
        print()  # newline after progress

        if result["success"]:
            print(f"  [ok] {len(result['models_pulled'])} component(s) ready\n")
            # Re-initialize kernel to pick up the new Ollama backend
            kernel = IntentKernel()
        else:
            for err in result["errors"]:
                print(f"  [!!] {err}")
            print()

    except Exception as exc:
        print(f"  Setup error: {exc}\n")

    return kernel


# ---------------------------------------------------------------------------
# CLI main loop — rich + prompt_toolkit
# ---------------------------------------------------------------------------

# Brand colour
_BRAND = "rgb(37,99,235)"

_LYNX = """\
      /\\_____/\\
     /  o   o  \\
    ( ==  ^  == )
     )         (
    (           )
   ( (  )   (  ) )
  (__(__)___(__)__)"""

_HELP_TABLE = [
    ("!help", "Show this help"),
    ("!setup", "Set up AI backend (Ollama or API key)"),
    ("!voice / !v", "Voice input (speak a task)"),
    ("!speak <text>", "Speak text aloud"),
    ("!status", "System status"),
    ("!cost", "Cost breakdown"),
    ("!history", "Task history"),
    ("!credentials", "Credential info"),
    ("!security", "Security stats"),
    ("exit", "Quit"),
]


def _get_prompt_session():
    """Create a prompt_toolkit session with history and autocomplete."""
    try:
        from prompt_toolkit import PromptSession
        from prompt_toolkit.history import FileHistory
        from prompt_toolkit.auto_suggest import AutoSuggestFromHistory
        from prompt_toolkit.completion import WordCompleter

        history_path = os.path.join(
            os.path.expanduser("~"), ".intentos", "cli_history"
        )
        commands = [
            "!help", "!setup", "!voice", "!v", "!speak", "!status",
            "!cost", "!history", "!credentials", "!security", "exit", "quit",
        ]
        completer = WordCompleter(commands, sentence=True)
        return PromptSession(
            history=FileHistory(history_path),
            auto_suggest=AutoSuggestFromHistory(),
            completer=completer,
            complete_while_typing=False,
        )
    except ImportError:
        return None


def main() -> None:
    import argparse
    import signal
    import threading

    parser = argparse.ArgumentParser(description="IntentOS Kernel")
    parser.add_argument("--headless", action="store_true",
                        help="Run in headless mode (API server only, no REPL)")
    parser.add_argument("--host", default="127.0.0.1",
                        help="API bridge host (default: 127.0.0.1)")
    parser.add_argument("--port", type=int, default=7891,
                        help="API bridge port (default: 7891)")
    args = parser.parse_args()

    _ensure_workspace()

    # Import rich (graceful fallback to plain print if missing)
    try:
        from rich.console import Console
        from rich.panel import Panel
        from rich.table import Table
        from rich.text import Text
        from rich.status import Status
        console = Console()
        HAS_RICH = True
    except ImportError:
        console = None
        HAS_RICH = False

    def cprint(*a, **kw):
        if console:
            console.print(*a, **kw)
        else:
            print(*a)

    if args.headless:
        kernel = IntentKernel()
        try:
            from core.api.server import APIBridge
            api = APIBridge(kernel=kernel)
            api.start(port=args.port)
            cprint(f"IntentOS headless — API on {args.host}:{args.port}")
        except Exception as e:
            cprint(f"API bridge failed: {e}")
            return
        shutdown = threading.Event()
        signal.signal(signal.SIGINT, lambda *_: shutdown.set())
        signal.signal(signal.SIGTERM, lambda *_: shutdown.set())
        shutdown.wait()
        return

    # First-run wizard
    try:
        from core.first_run import FirstRunWizard
        wizard = FirstRunWizard()
        if wizard.is_first_run():
            cprint()
            cprint(f"  [{_BRAND}]{_LYNX}[/{_BRAND}]" if HAS_RICH else _LYNX)
            cprint()
            cprint("  [bold]Welcome to IntentOS[/bold]" if HAS_RICH else "  Welcome to IntentOS")
            cprint("  [dim]Your computer, finally on your side.[/dim]" if HAS_RICH else "  Your computer, finally on your side.")
            cprint()
            wizard.run(skip_prompts=False)
    except Exception as e:
        cprint(f"  Setup note: {e}")
        cprint("  Continuing with defaults...\n")

    kernel = IntentKernel()

    # Start API bridge in background
    try:
        from core.api.server import APIBridge
        api = APIBridge(kernel=kernel)
        api.start(port=7891)
        api_status = "API on :7891"
    except Exception:
        api_status = "API offline"

    config = kernel.llm.get_config()
    model = config.get('local_model', 'N/A')
    mode = config.get('privacy_mode', 'N/A')
    has_local = kernel.llm._router._local_backend is not None
    has_cloud = kernel.llm._router._cloud_backend is not None

    if has_local:
        model_display = model
    elif has_cloud:
        model_display = config.get('cloud_model', 'cloud')
    else:
        model_display = "none"

    # Banner
    cprint()
    if HAS_RICH:
        cprint(f"  [{_BRAND}]{_LYNX}[/{_BRAND}]")
        cprint()
        cprint(f"  [bold]IntentOS[/bold] v{IntentKernel.VERSION}")
        cprint(f"  [dim]Language is the interface. The file never leaves.[/dim]")
        cprint()
        # Status line
        status_parts = []
        backend_tag = "[green]local[/green]" if has_local else "[yellow]cloud[/yellow]" if has_cloud else "[red]none[/red]"
        status_parts.append(f"Model: [bold]{model_display}[/bold] ({backend_tag})")
        status_parts.append(f"Mode: {mode}")
        status_parts.append(api_status)
        cprint(f"  {' | '.join(status_parts)}")
    else:
        print(f"  {_LYNX}")
        print(f"\n  IntentOS v{IntentKernel.VERSION}")
        print(f"  Model: {model_display} | Mode: {mode} | {api_status}")

    if not has_local and not has_cloud:
        kernel = _interactive_setup(kernel, model)
        has_local = kernel.llm._router._local_backend is not None
        has_cloud = kernel.llm._router._cloud_backend is not None
        if has_local:
            model_display = model
        elif has_cloud:
            model_display = config.get('cloud_model', 'cloud')

        if has_local or has_cloud:
            cprint(f"\n  Model: {model_display} | Ready")
        else:
            cprint("\n  No backend configured. Type '!setup' to try again.")
    elif not has_local:
        cprint(f"  [yellow]Local model not available. Using cloud API only.[/yellow]" if HAS_RICH
               else "  Local model not available. Using cloud API only.")

    cprint(f"\n  Type a task in plain English, or [bold]!help[/bold] for commands.\n" if HAS_RICH
           else "\n  Type a task in plain English, or !help for commands.\n")

    # Set up prompt_toolkit session (falls back to input() if unavailable)
    session = _get_prompt_session()

    def get_input() -> str:
        if session:
            from prompt_toolkit.formatted_text import HTML
            return session.prompt(HTML('<b>intentos</b>&gt; ')).strip()
        return input("intentos> ").strip()

    # REPL loop
    while True:
        try:
            user_input = get_input()
        except (EOFError, KeyboardInterrupt):
            cprint("\n  Goodbye.")
            break

        if not user_input:
            continue

        if user_input.lower() in ("exit", "quit"):
            stats = kernel.llm.get_stats()
            cprint()
            if HAS_RICH:
                cprint(f"  [dim]Session: {stats['total_calls']} calls, "
                       f"{stats['total_tokens']} tokens, "
                       f"${stats['total_cost_usd']:.4f}[/dim]")
            else:
                print(f"  Session: {stats['total_calls']} calls, "
                      f"{stats['total_tokens']} tokens, "
                      f"${stats['total_cost_usd']:.4f}")
            break

        # Handle CLI subcommands
        if user_input.startswith("!"):
            cmd_text = user_input[1:].strip().lower()

            if cmd_text in ("voice", "v"):
                try:
                    from core.voice.stt import VoiceInput
                    vi = VoiceInput()
                    if vi.is_available():
                        cprint("  Listening... (speak now, 5 seconds)")
                        result = vi.listen_and_transcribe(duration=5)
                        if result and result.text:
                            cprint(f'  Heard: "{result.text}"')
                            user_input = result.text
                        else:
                            cprint("  Couldn't understand that. Try again or type your task.")
                            continue
                    else:
                        cprint("  Voice input not available. Install: pip install SpeechRecognition pyaudio")
                        continue
                except Exception as e:
                    cprint(f"  Voice error: {e}")
                    continue

            elif cmd_text == "setup":
                cfg = kernel.llm.get_config()
                m = cfg.get('local_model', 'gemma4:e4b')
                kernel = _interactive_setup(kernel, m)
                continue

            elif cmd_text == "help":
                if HAS_RICH:
                    table = Table(show_header=False, box=None, padding=(0, 2))
                    table.add_column(style="bold cyan", no_wrap=True)
                    table.add_column(style="dim")
                    for cmd, desc in _HELP_TABLE:
                        table.add_row(cmd, desc)
                    cprint()
                    cprint(table)
                    cprint()
                else:
                    for cmd, desc in _HELP_TABLE:
                        print(f"  {cmd:20s} {desc}")
                continue

            else:
                output = kernel.handle_command(cmd_text)
                cprint(output)
                continue

        # Execute task with spinner
        if HAS_RICH:
            with console.status("[bold blue]Thinking...", spinner="dots"):
                result = kernel.process_task(user_input)
        else:
            result = kernel.process_task(user_input)

        kernel.display_result(result, console=console if HAS_RICH else None)


if __name__ == "__main__":
    main()
